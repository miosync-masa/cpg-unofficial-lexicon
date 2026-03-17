import json, numpy as np, os, csv, warnings
warnings.filterwarnings('ignore')
from scipy.optimize import differential_evolution, minimize
import pandas as pd
from dotenv import load_dotenv

BASE = os.path.expanduser('~/multi-agent-shogun')
RESULTS = os.path.join(BASE, 'results')

# Load API key from .env
load_dotenv(os.path.join(BASE, '.env'))
api_key = os.environ.get('OPENAI_API_KEY', '')
if not api_key:
    print("ERROR: OPENAI_API_KEY not found in .env")
    exit(1)

import openai
client = openai.OpenAI(api_key=api_key)
MODEL = 'text-embedding-3-large'  # 3072 dimensions!

print("=" * 70)
print("CPG FULL PIPELINE: 217 items x text-embedding-3-large (3072d)")
print("=" * 70)

# ================================================================
# STEP 1: Load both datasets
# ================================================================

# A) Shogun 114 items (from CSV)
shogun_rows = []
csv_path = os.path.join(RESULTS, 'cpg_all_items_for_embedding.csv')
with open(csv_path, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        shogun_rows.append(row)
print("Shogun items: %d" % len(shogun_rows))

# Fill r,e,s from YAMLs (same logic as before)
import yaml
yaml_lookup = {}
for fname in os.listdir(RESULTS):
    if fname.startswith('ashigaru') and fname.endswith('.yaml'):
        with open(os.path.join(RESULTS, fname)) as f:
            y = yaml.safe_load(f)
        for item in y.get('items', []):
            iid = item.get('id', '')
            r = item.get('x_r', item.get('r', None))
            e = item.get('x_e', item.get('e', None))
            s = item.get('x_s', item.get('s', None))
            xt = item.get('x_type', None)
            if xt and isinstance(xt, list) and len(xt) == 3:
                r, e, s = xt[0], xt[1], xt[2]
            if r is not None and e is not None and s is not None:
                yaml_lookup[iid] = (float(r), float(e), float(s))

import openpyxl# B) Torami 103 items (from xlsx)
xlsx_paths = [
    os.path.expanduser('~/multi-agent-shogun/cpg_coding_table_v3.xlsx'),
    os.path.expanduser('~/Desktop/cpg_coding_table_v3.xlsx'),
    os.path.expanduser('~/Downloads/cpg_coding_table_v3.xlsx'),
]
wb = None
for p in xlsx_paths:
    if os.path.exists(p):
        wb = openpyxl.load_workbook(p, data_only=True)
        print("Found xlsx at: %s" % p)
        break

# Also check in results and project root
for p in [os.path.join(RESULTS, 'cpg_coding_table_v3.xlsx'),
          os.path.join(BASE, 'cpg_coding_table_v3.xlsx')]:
    if os.path.exists(p) and wb is None:
        wb = openpyxl.load_workbook(p, data_only=True)
        print("Found xlsx at: %s" % p)
        break

torami_rows = []
if wb:
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None or not isinstance(row[0], int):
            continue
        try:
            r_val = float(row[7]) if row[7] is not None else None
            e_val = float(row[8]) if row[8] is not None else None
            s_val = float(row[9]) if row[9] is not None else None
        except (ValueError, TypeError):
            continue

        host = str(row[4] or row[3] or '')  # Romanized or Original
        target = str(row[5] or '')
        vector = str(row[10] or '')
        z_reg = row[11]
        z_vel = str(row[12] or '')
        exp = str(row[1] or '')

        torami_rows.append({
            'id': 'T_%03d' % row[0],
            'experiment': exp,
            'language': str(row[2] or ''),
            'host_word': str(row[3] or ''),
            'host_romanized': host,
            'covert_target': target,
            'r': r_val, 'e': e_val, 's': s_val,
            'vector': vector,
            'z_register': int(z_reg) if z_reg else 2,
            'z_velocity': z_vel,
            'source': 'torami_v3',
        })
    print("Torami items: %d" % len(torami_rows))
else:
    print("WARNING: cpg_coding_table_v3.xlsx not found! Running with shogun data only.")

# C) Build unified list
all_items = []

# Shogun items
for row in shogun_rows:
    iid = row['id']
    r = float(row['r']) if row.get('r') else None
    e = float(row['e']) if row.get('e') else None
    s = float(row['s']) if row.get('s') else None
    if (r is None or e is None or s is None) and iid in yaml_lookup:
        r, e, s = yaml_lookup[iid]
    host = row.get('host_romanized', '') or row.get('host_word', '')
    target = row.get('covert_target', '')
    all_items.append({
        'id': iid, 'host': host.strip(), 'target': target.strip(),
        'r': r, 'e': e, 's': s,
        'vector': row.get('vector', ''),
        'z_register': int(row['z_register']) if row.get('z_register') else 2,
        'z_velocity': row.get('z_velocity', ''),
        'source': 'shogun',
    })

# Torami items
for row in torami_rows:
    all_items.append({
        'id': row['id'], 'host': row['host_romanized'].strip(),
        'target': row['covert_target'].strip(),
        'r': row['r'], 'e': row['e'], 's': row['s'],
        'vector': row['vector'],
        'z_register': row['z_register'],
        'z_velocity': row['z_velocity'],
        'source': 'torami_v3',
    })

print("\nTotal unified items: %d (shogun=%d + torami=%d)" %
      (len(all_items), len(shogun_rows), len(torami_rows)))

# Filter valid items
valid = [i for i in all_items if i['r'] is not None and i['host'] and i['target']]
print("Valid for embedding: %d" % len(valid))

# ================================================================
# STEP 2: Embedding with text-embedding-3-large
# ================================================================
host_texts = [v['host'] for v in valid]
target_texts = [v['target'] for v in valid]

def batch_embed(texts, label):
    all_embs = []
    batch_size = 100
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i+batch_size]
        print("  %s batch %d (%d items)..." % (label, i//batch_size+1, len(batch)))
        resp = client.embeddings.create(model=MODEL, input=batch)
        for d in resp.data:
            all_embs.append(d.embedding)
    return all_embs

print("\nEmbedding with %s (3072 dimensions)..." % MODEL)
print("Host words:")
host_embs = batch_embed(host_texts, "host")
print("Covert targets:")
target_embs = batch_embed(target_texts, "target")

# Cosine similarity
def cosine_sim(a, b):
    a, b = np.array(a), np.array(b)
    return float(np.dot(a, b) / (np.linalg.norm(a) * np.linalg.norm(b)))

for i, v in enumerate(valid):
    v['sim'] = cosine_sim(host_embs[i], target_embs[i])

print("\nEmbedding complete!")
print("Mean sim: %.4f, Std: %.4f" % (np.mean([v['sim'] for v in valid]),
                                       np.std([v['sim'] for v in valid])))

# ================================================================
# STEP 3: Build DataFrame
# ================================================================
rows = []
for v in valid:
    region = v['id'].split('_')[0] if v['source'] == 'shogun' else v.get('experiment', 'T').split('-')[0] if v['source'] == 'torami_v3' else 'UNK'
    if v['source'] == 'torami_v3':
        exp = v.get('experiment', '') if 'experiment' in v else ''
        # Map torami experiments to region codes
        region = 'T_' + region  # prefix with T_ to distinguish
    vecs = [x.strip() for x in v.get('vector','').split('+') if x.strip() and x.strip() != 'Formal?']
    rows.append({
        'id': v['id'], 'sim': v['sim'],
        'R': v['r'], 'E': v['e'], 'S': v['s'],
        'Z': v['z_register'],
        'region': region, 'source': v['source'],
        'has_Phonetic': 1 if 'Phonetic' in vecs else 0,
        'has_Formal': 1 if 'Formal' in vecs else 0,
        'has_Metonymic': 1 if 'Metonymic' in vecs else 0,
        'has_Behavioral': 1 if 'Behavioral' in vecs else 0,
        'has_Structural': 1 if 'Structural' in vecs else 0,
        'has_Morphological': 1 if 'Morphological' in vecs else 0,
        'has_Chromatic': 1 if 'Chromatic' in vecs else 0,
    })

df = pd.DataFrame(rows)
n = len(df)
Y = df['sim'].values
R = df['R'].values
E = df['E'].values
S = df['S'].values
Z = df['Z'].values
P = df['has_Phonetic'].values
F = df['has_Formal'].values

SS_tot = np.sum((Y - np.mean(Y))**2)

print("\nDataFrame: %d rows" % n)
print("Sources:", df['source'].value_counts().to_dict())
print("Regions:", df['region'].nunique(), "unique")

# ================================================================
# MODEL 1: OLS
# ================================================================
X1 = np.column_stack([np.ones(n), R, E, Z])
b1 = np.linalg.lstsq(X1, Y, rcond=None)[0]
R2_ols = 1 - np.sum((Y - X1@b1)**2) / SS_tot
MSE1 = np.sum((Y - X1@b1)**2) / (n-4)
var1 = MSE1 * np.linalg.inv(X1.T @ X1)
se1 = np.sqrt(np.diag(var1))
t1 = b1 / se1

print("\n" + "=" * 70)
print("MODEL 1: OLS sim ~ R + E + Z  (N=%d, 3072d LARGE)" % n)
print("=" * 70)
print("R2=%.4f  AdjR2=%.4f" % (R2_ols, 1-(1-R2_ols)*(n-1)/(n-4)))
for nm, b, se, t in zip(['const','R','E','Z'], b1, se1, t1):
    p = 2*(1-__import__('scipy').stats.t.cdf(abs(t), n-4))
    sig = '***' if p<0.001 else '**' if p<0.01 else '*' if p<0.05 else '.' if p<0.1 else ''
    print("  %-10s b=%8.4f se=%7.4f t=%6.2f p=%.4f %s" % (nm, b, se, t, p, sig))

# ================================================================
# MODEL 2: HLM
# ================================================================
from statsmodels.regression.mixed_linear_model import MixedLM
import statsmodels.api as sm

X_hlm = df[['R','E','Z']].copy()
X_hlm = sm.add_constant(X_hlm)
hlm = MixedLM(df['sim'], X_hlm, groups=df['region']).fit(reml=True)

var_re = float(hlm.cov_re.iloc[0,0])
icc = var_re / (var_re + hlm.scale)

print("\n" + "=" * 70)
print("MODEL 2: HLM + (1|region)  (N=%d)" % n)
print("=" * 70)
print(hlm.summary().tables[1])
print("ICC = %.4f (%.1f%% between-region)" % (icc, icc*100))

# HLM Full with vectors
vec_cols = ['has_Phonetic','has_Formal','has_Metonymic','has_Behavioral','has_Structural','has_Morphological']
X_d = df[['R','E','Z'] + vec_cols].copy()
X_d = sm.add_constant(X_d)
hlm_d = MixedLM(df['sim'], X_d, groups=df['region']).fit(reml=True)

print("\n" + "=" * 70)
print("MODEL 2b: HLM Full + Vectors")
print("=" * 70)
print(hlm_d.summary().tables[1])

# ================================================================
# MODEL 3: Gate x Heaviside (PCC x SCC)
# ================================================================
def gc_step(p, R, E, Z, S, P, F):
    a0,a1,a2,a3,a4,a5,alpha,K = p
    gate = a0 + a1*R + a2*E + a3*Z + a4*P + a5*F
    ch = np.where(S > K, 1.0 - alpha, 1.0)
    return gate * ch

def loss_step(p):
    return np.sum((Y - gc_step(p, R, E, Z, S, P, F))**2)

bounds_step = [(-0.5,1),(-0.5,0.5),(-0.5,0.5),(-0.1,0.2),(-0.3,0.3),(-0.3,0.3),(0.01,0.99),(0.01,0.99)]
res_h = differential_evolution(loss_step, bounds_step, seed=42, maxiter=2000, tol=1e-12)
ph = res_h.x
Yh = gc_step(ph, R, E, Z, S, P, F)
R2_h = 1 - np.sum((Y - Yh)**2) / SS_tot
R2_adj_h = 1 - (1-R2_h)*(n-1)/(n-len(ph)-1)

print("\n" + "=" * 70)
print("MODEL 3: Gate x Heaviside (N=%d, 3072d LARGE)" % n)
print("=" * 70)
print("R2=%.4f  AdjR2=%.4f" % (R2_h, R2_adj_h))
print("Gate: a0=%.4f R=%.4f E=%.4f Z=%.4f Phonetic=%.4f Formal=%.4f" %
      (ph[0],ph[1],ph[2],ph[3],ph[4],ph[5]))
print("Channel: alpha=%.4f  K_crit=%.4f" % (ph[6], ph[7]))
print("  S < %.4f: Channel=1.000 (no emergence)" % ph[7])
print("  S > %.4f: Channel=%.3f (emergence ON)" % (ph[7], 1-ph[6]))

# ================================================================
# MODEL 4: Region-centered Heaviside
# ================================================================
region_means = df.groupby('region')['sim'].mean()
grand_mean = Y.mean()
offsets = df['region'].map(region_means).values - grand_mean
Yc = Y - offsets

def loss_step_c(p):
    return np.sum((Yc - gc_step(p, R, E, Z, S, P, F))**2)

res_rc = differential_evolution(loss_step_c, bounds_step, seed=42, maxiter=2000, tol=1e-12)
prc = res_rc.x
Yrc = gc_step(prc, R, E, Z, S, P, F) + offsets
R2_rc = 1 - np.sum((Y - Yrc)**2) / SS_tot

print("\n" + "=" * 70)
print("MODEL 4: Region-centered Heaviside (N=%d, 3072d LARGE)" % n)
print("=" * 70)
print("R2=%.4f" % R2_rc)
print("Channel: alpha=%.4f  K_crit=%.4f" % (prc[6], prc[7]))

# ================================================================
# S ANALYSIS
# ================================================================
print("\n" + "=" * 70)
print("S-COMPONENT ANALYSIS (3072d LARGE)")
print("=" * 70)
s0 = df[df['S']==0]; snz = df[df['S']>0]; sd = df[df['S']>=0.4]
print("S=0    (n=%d): sim=%.4f" % (len(s0), s0['sim'].mean()))
print("S>0    (n=%d): sim=%.4f" % (len(snz), snz['sim'].mean()))
print("S>=0.4 (n=%d): sim=%.4f" % (len(sd), sd['sim'].mean()))
from scipy.stats import ttest_ind
if len(s0)>0 and len(snz)>0:
    t,p = ttest_ind(s0['sim'], snz['sim'], equal_var=False)
    d = (s0['sim'].mean()-snz['sim'].mean()) / np.sqrt((s0['sim'].std()**2+snz['sim'].std()**2)/2)
    print("Welch t=%.3f  p=%.4f  Cohen d=%.3f" % (t, p, d))

# ================================================================
# TYPE ANALYSIS
# ================================================================
print("\n" + "=" * 70)
print("TYPE ANALYSIS (3072d LARGE)")
print("=" * 70)
rd = df[df['R']>=0.6]; ed = df[df['E']>=0.6]
print("R-dom (n=%d): sim=%.4f" % (len(rd), rd['sim'].mean()))
print("E-dom (n=%d): sim=%.4f" % (len(ed), ed['sim'].mean()))
if len(sd)>0:
    print("S-dom (n=%d): sim=%.4f" % (len(sd), sd['sim'].mean()))

# ================================================================
# REGION ANALYSIS
# ================================================================
print("\n" + "=" * 70)
print("REGION ANALYSIS (3072d LARGE)")
print("=" * 70)
for reg, grp in df.groupby('region'):
    print("%8s (n=%2d): sim=%.4f +/- %.4f" % (reg, len(grp), grp['sim'].mean(), grp['sim'].std()))

# ================================================================
# COMPARISON
# ================================================================
print("\n" + "=" * 70)
print("FINAL MODEL COMPARISON (text-embedding-3-large, N=%d)" % n)
print("=" * 70)
print("%-45s  R2      AdjR2" % "Model")
print("-" * 70)
print("%-45s  %.4f  %.4f" % ("1. OLS: sim ~ R+E+Z", R2_ols, 1-(1-R2_ols)*(n-1)/(n-4)))
print("%-45s  ICC=%.4f" % ("2. HLM: + (1|region)", icc))
print("%-45s  %.4f  %.4f" % ("3. Gate x Heaviside (PCC*SCC)", R2_h, R2_adj_h))
print("%-45s  %.4f" % ("4. Region + Heaviside", R2_rc))
print("\nsmall(1536d) vs LARGE(3072d) comparison:")
print("  OLS:       0.0494 -> %.4f" % R2_ols)
print("  Heaviside: 0.1368 -> %.4f" % R2_h)
print("  Region+H:  0.3375 -> %.4f" % R2_rc)

# ================================================================
# SAVE
# ================================================================
results = {
    'model': MODEL, 'dimensions': 3072, 'n_total': n,
    'n_shogun': len([v for v in valid if v.get('source')=='shogun']),
    'n_torami': len([v for v in valid if v.get('source')=='torami_v3']),
    'ols': {'R2': round(R2_ols,4)},
    'hlm': {'ICC': round(icc,4)},
    'heaviside': {'R2': round(R2_h,4), 'K_crit': round(float(ph[7]),4), 'alpha': round(float(ph[6]),4)},
    'region_heaviside': {'R2': round(R2_rc,4)},
    'mean_sim': round(float(np.mean(Y)),4),
    'std_sim': round(float(np.std(Y)),4),
}
with open(os.path.join(RESULTS, 'cpg_large_embedding_results.json'), 'w') as f:
    json.dump(results, f, indent=2)

# Save embeddings
np.savez_compressed(os.path.join(RESULTS, 'cpg_large_embeddings.npz'),
    host=np.array(host_embs), target=np.array(target_embs),
    ids=np.array([v['id'] for v in valid]),
    sims=np.array([v['sim'] for v in valid]))

print("\nSaved to cpg_large_embedding_results.json + cpg_large_embeddings.npz")
print("\n" + "=" * 70)
print("DONE!")
print("=" * 70)
